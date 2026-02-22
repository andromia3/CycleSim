"""Full render pipeline test -- catches every error the callback would hit."""
import warnings
import traceback
import sys
sys.path.insert(0, ".")
warnings.filterwarnings("ignore")

from ui.state import get_or_run, run_sensitivity
from cyclesim.defaults import *

params = {
    "n_paths": 1000, "n_years": 25, "seed": 42, "ar2_mode": "auto",
    "n_strategies": 2,
    "cycle_period": MARKET_DEFAULTS["cycle_period_years"],
    "min_lr": MARKET_DEFAULTS["min_loss_ratio"],
    "max_lr": MARKET_DEFAULTS["max_loss_ratio"],
    "long_run_lr": MARKET_DEFAULTS["long_run_loss_ratio"],
    "phi_1": 1.1, "phi_2": -0.45, "sigma_epsilon": 0.10,
    "shock_prob": MARKET_DEFAULTS["shock_prob"],
    "shock_magnitude": MARKET_DEFAULTS["shock_magnitude_std"],
    "attritional_mean": LOSS_DEFAULTS["attritional_mean"],
    "cycle_sensitivity": LOSS_DEFAULTS["cycle_sensitivity"],
    "tail_sensitivity": LOSS_DEFAULTS["tail_thickness_sensitivity"],
    "large_freq": LOSS_DEFAULTS["large_loss_frequency"],
    "cat_freq": LOSS_DEFAULTS["cat_frequency"],
    "cat_severity": LOSS_DEFAULTS["cat_severity_mean"],
}

# Add insurer params for all 6 prefixes
for pfx in STRATEGY_PREFIXES:
    preset = INSURER_PRESETS[pfx]
    params.update({
        f"{pfx}_name": preset["name"],
        f"{pfx}_gwp": preset["initial_gwp"],
        f"{pfx}_expense": preset["expense_ratio"],
        f"{pfx}_cession": preset["base_cession_pct"],
        f"{pfx}_inv_return": preset["investment_return"],
        f"{pfx}_expected_lr": preset["expected_lr"],
        f"{pfx}_growth": preset["growth_rate_when_profitable"],
        f"{pfx}_shrink": preset["shrink_rate_when_unprofitable"],
        f"{pfx}_max_growth": preset["max_gwp_growth_pa"],
        f"{pfx}_max_shrink": preset["max_gwp_shrink_pa"],
        f"{pfx}_cess_sens": preset["cession_cycle_sensitivity"],
        f"{pfx}_min_cess": preset["min_cession_pct"],
        f"{pfx}_max_cess": preset["max_cession_pct"],
        f"{pfx}_adv_sel": preset["adverse_selection_sensitivity"],
        f"{pfx}_cap_ratio": preset["capital_ratio"],
        f"{pfx}_coc": preset["cost_of_capital"],
        # Signal weights (defaults)
        f"{pfx}_sw_own_lr": SIGNAL_WEIGHT_DEFAULTS["own_lr"],
        f"{pfx}_sw_market_lr": SIGNAL_WEIGHT_DEFAULTS["market_lr"],
        f"{pfx}_sw_rate_adequacy": SIGNAL_WEIGHT_DEFAULTS["rate_adequacy"],
        f"{pfx}_sw_rate_change": SIGNAL_WEIGHT_DEFAULTS["rate_change"],
        f"{pfx}_sw_capital": SIGNAL_WEIGHT_DEFAULTS["capital"],
    })

fails = []

results = get_or_run(params)
summaries = results.summaries
ins_list = results.insurers
ny = results.config.n_years
n = len(summaries)
names = [results.config.insurers[i].name for i in range(n)]

from ui.exhibits import (
    kpi_row, convergence_indicator, executive_summary, summary_table,
    regime_performance_table, return_period_table, stress_scenario_cards,
    worst_paths_table, drilldown_table, strategy_profile_card,
)
from ui.charts import (
    market_cycle_chart, radar_chart, attribution_chart, fan_chart,
    comparison_fan_chart, distribution_chart, ruin_over_time_chart,
    rorac_scatter, efficiency_frontier, drawdown_chart,
    win_probability_chart, cycle_timing_chart, solvency_comparison_chart,
    cr_regime_heatmap, yearly_profit_waterfall, cumulative_profit_buildup,
    capital_model_preview, sensitivity_tornado_from_rows,
    regime_performance_data, worst_paths_data, single_path_chart,
    # Pass 26
    historical_cr_chart, historical_cumulative_chart, counterfactual_waterfall,
    regime_forecast_chart, market_clock_chart, tail_decomposition_chart,
    strategy_dna_radar,
    STRATEGY_COLORS, STRATEGY_COLORS_LIGHT, STRATEGY_COLORS_MED,
    COLOR_A, COLOR_B,
)
from ui.exhibits import (
    backtest_summary_cards, backtest_year_table, counterfactual_verdict,
    regime_outlook_table, tail_decomposition_table,
    data_provenance_badge,
)
from ui.app import _insurer_tab, _market_rate_chart, _regime_bar, _market_stats_bar
from cyclesim.metrics import compute_return_period_table, compute_tail_decomposition
from cyclesim.market import compute_regime_forecast
from cyclesim.historical import replay_historical, counterfactual_decomposition, LLOYDS_HISTORICAL
import numpy as np

rp_rows = regime_performance_data(results.market, ins_list, names)

# Pass 26 pre-computations
config = results.config
loss_params = config.loss_params
bt_list = [replay_historical(config.insurers[i], loss_params) for i in range(n)]
cf_decomp = counterfactual_decomposition(bt_list[0], bt_list[1], config.insurers[0], config.insurers[1], loss_params)
decomp_list = [compute_tail_decomposition(ins_list[i], config.insurers[i]) for i in range(n)]
trans_matrix = results.market["params"].regime_transition_matrix
regime_counts = np.bincount(results.market["regime"][:, 0], minlength=4)
current_regime = int(np.argmax(regime_counts))
forecast = compute_regime_forecast(trans_matrix, current_regime, n_years=10)

# Build regime perf dicts for outlook table
regime_perfs = []
for i in range(n):
    regime_perfs.append({
        r["regime"].lower(): {"cr": r["strategies"][i]["cr"], "rorac": r["strategies"][i]["rorac"]}
        for r in rp_rows
    })

tests = {
    # Exhibits (list-based signatures)
    "kpi_row": lambda: kpi_row(summaries, names, results.elapsed_seconds),
    "convergence": lambda: convergence_indicator(
        [s["through_cycle_rorac_dist"] for s in summaries],
        [s["cumulative_profit_dist"] for s in summaries], 1000),
    "exec_summary": lambda: executive_summary(summaries, names, 1000, ny, results.elapsed_seconds, rp_rows),
    "summary_table": lambda: summary_table(summaries, names),
    "regime_perf": lambda: regime_performance_table(rp_rows, names),
    "strategy_a": lambda: strategy_profile_card("Test", 0.08, -0.05, 0.15, -0.20, 0.55, 0.03, 0.23, 0.10, 0.45),
    "return_period": lambda: return_period_table(
        compute_return_period_table(results.insurers, results.config.insurers), names),
    "stress_cards": lambda: stress_scenario_cards(
        compute_return_period_table(results.insurers, results.config.insurers), names),
    "worst_paths": lambda: worst_paths_table(worst_paths_data(results, 10), names),
    "drilldown_tbl": lambda: drilldown_table(results.market, ins_list, 0, ny, names),
    # Charts
    "market_cycle": lambda: market_cycle_chart(results.market, ny),
    "radar": lambda: radar_chart(summaries, names),
    "attribution": lambda: attribution_chart([s["attribution"] for s in summaries], names),
    "fan_gwp": lambda: fan_chart(summaries[0]["percentile_bands"]["gwp"], summaries[0]["yearly_means"]["gwp"], ny, "GWP", "GBP"),
    "compare_gwp": lambda: comparison_fan_chart(
        [s["percentile_bands"]["gwp"] for s in summaries],
        [s["yearly_means"]["gwp"] for s in summaries],
        ny, "GWP", "GBP", names),
    "dist_profit": lambda: distribution_chart(
        [s["cumulative_profit_dist"] for s in summaries], "P", "GBP", names),
    "ruin": lambda: ruin_over_time_chart([s["ruin_prob_by_year"] for s in summaries], ny, names),
    "rorac_scatter": lambda: rorac_scatter([s["through_cycle_rorac_dist"] for s in summaries], names),
    "efficiency": lambda: efficiency_frontier(summaries, names),
    "drawdown": lambda: drawdown_chart([d["capital"] for d in ins_list], ny, names),
    "win_prob": lambda: win_probability_chart(
        [d["cumulative_profit"] for d in ins_list], ny, names),
    "cycle_timing": lambda: cycle_timing_chart(results.market, ins_list, names),
    "solvency": lambda: solvency_comparison_chart(
        [d["solvency_ratio"] for d in ins_list], ny, names),
    "cr_heatmap_a": lambda: cr_regime_heatmap(results.market, ins_list[0], ny, names[0], STRATEGY_COLORS[0]),
    "cr_heatmap_b": lambda: cr_regime_heatmap(results.market, ins_list[1], ny, names[1], STRATEGY_COLORS[1]),
    "waterfall_a": lambda: yearly_profit_waterfall(ins_list[0], ny, names[0], STRATEGY_COLORS[0]),
    "waterfall_b": lambda: yearly_profit_waterfall(ins_list[1], ny, names[1], STRATEGY_COLORS[1]),
    "buildup": lambda: cumulative_profit_buildup(ins_list, ny, names),
    "drilldown_chart": lambda: single_path_chart(results.market, ins_list, 0, ny, names),
    "rate_chart": lambda: _market_rate_chart(results.market, ny),
    "regime_bar": lambda: _regime_bar(results.market, ny),
    "market_stats": lambda: _market_stats_bar(results.market, ny),
    # Insurer tabs (the big composite ones)
    "insurer_tab_a": lambda: _insurer_tab(summaries[0], ny, names[0],
        STRATEGY_COLORS[0], STRATEGY_COLORS_LIGHT[0], STRATEGY_COLORS_MED[0],
        results.config.insurers[0], results.market, ins_list[0]),
    "insurer_tab_b": lambda: _insurer_tab(summaries[1], ny, names[1],
        STRATEGY_COLORS[1], STRATEGY_COLORS_LIGHT[1], STRATEGY_COLORS_MED[1],
        results.config.insurers[1], results.market, ins_list[1]),
    # Sensitivity
    "sensitivity": lambda: sensitivity_tornado_from_rows(run_sensitivity(params, 200), names),
    # Pass 26: Backtest
    "backtest_cards": lambda: backtest_summary_cards(bt_list, names),
    "backtest_year_tbl": lambda: backtest_year_table(bt_list, LLOYDS_HISTORICAL, names),
    "backtest_cr_chart": lambda: historical_cr_chart(bt_list, LLOYDS_HISTORICAL, names),
    "backtest_cumulative": lambda: historical_cumulative_chart(bt_list, names),
    "counterfactual_verdict": lambda: counterfactual_verdict(cf_decomp, names[0], names[1]),
    "counterfactual_waterfall": lambda: counterfactual_waterfall(cf_decomp),
    # Pass 26: Regime Forecast
    "regime_forecast": lambda: regime_forecast_chart(forecast, current_regime),
    "market_clock": lambda: market_clock_chart(trans_matrix, current_regime),
    "regime_outlook_tbl": lambda: regime_outlook_table(forecast, regime_perfs, names),
    # Pass 26: Tail Decomposition
    "tail_decomp_chart": lambda: tail_decomposition_chart(decomp_list, names),
    "tail_decomp_table": lambda: tail_decomposition_table(decomp_list, names),
    # Pass 26: Provenance
    "provenance_synthetic": lambda: data_provenance_badge(has_capital_model=False),
    "provenance_uploaded": lambda: data_provenance_badge(has_capital_model=True),
    # Pass 26: DNA Radar
    "strategy_dna": lambda: strategy_dna_radar(
        {k: getattr(config.insurers[0], k) for k in [
            "growth_rate_when_profitable", "shrink_rate_when_unprofitable",
            "max_gwp_growth_pa", "base_cession_pct", "cession_cycle_sensitivity",
            "expected_lr", "max_gwp_shrink_pa", "adverse_selection_sensitivity"]},
        {k: getattr(config.insurers[1], k) for k in [
            "growth_rate_when_profitable", "shrink_rate_when_unprofitable",
            "max_gwp_growth_pa", "base_cession_pct", "cession_cycle_sensitivity",
            "expected_lr", "max_gwp_shrink_pa", "adverse_selection_sensitivity"]},
    ),
}

for name, fn in tests.items():
    try:
        fn()
        print(f"  OK: {name}")
    except Exception as e:
        fails.append(name)
        print(f"  FAIL: {name}")
        traceback.print_exc()

# ---- N=3 render tests ----
print(f"\n{'='*50}")
print("N=3 STRATEGY TESTS")
params_3 = {**params, "n_strategies": 3}
results_3 = get_or_run(params_3)
n3 = 3
summaries_3 = results_3.summaries[:n3]
ins_list_3 = results_3.insurers[:n3]
names_3 = [results_3.config.insurers[i].name for i in range(n3)]

tests_3 = {
    "n3_kpi_row": lambda: kpi_row(summaries_3, names_3, results_3.elapsed_seconds),
    "n3_summary_table": lambda: summary_table(summaries_3, names_3),
    "n3_radar": lambda: radar_chart(summaries_3, names_3),
    "n3_rorac_scatter": lambda: rorac_scatter(
        [s["through_cycle_rorac_dist"] for s in summaries_3], names_3),
    "n3_drawdown": lambda: drawdown_chart(
        [d["capital"] for d in ins_list_3], results_3.config.n_years, names_3),
    "n3_win_prob": lambda: win_probability_chart(
        [d["cumulative_profit"] for d in ins_list_3], results_3.config.n_years, names_3),
    "n3_buildup": lambda: cumulative_profit_buildup(
        ins_list_3, results_3.config.n_years, names_3),
    "n3_efficiency": lambda: efficiency_frontier(summaries_3, names_3),
    "n3_drilldown": lambda: single_path_chart(
        results_3.market, ins_list_3, 0, results_3.config.n_years, names_3),
    "n3_sensitivity": lambda: sensitivity_tornado_from_rows(
        run_sensitivity(params_3, 200), names_3),
    "n3_backtest": lambda: backtest_summary_cards(
        [replay_historical(results_3.config.insurers[i], results_3.config.loss_params) for i in range(n3)],
        names_3),
}

for name, fn in tests_3.items():
    try:
        fn()
        print(f"  OK: {name}")
    except Exception as e:
        fails.append(name)
        print(f"  FAIL: {name}")
        traceback.print_exc()

# ---- N=6 (max strategies) render tests ----
print(f"\n{'='*50}")
print("N=6 STRATEGY TESTS")
params_6 = {**params, "n_strategies": 6}
results_6 = get_or_run(params_6)
n6 = 6
summaries_6 = results_6.summaries[:n6]
ins_list_6 = results_6.insurers[:n6]
names_6 = [results_6.config.insurers[i].name for i in range(n6)]

tests_6 = {
    "n6_summary_table": lambda: summary_table(summaries_6, names_6),
    "n6_radar": lambda: radar_chart(summaries_6, names_6),
    "n6_rorac_scatter": lambda: rorac_scatter(
        [s["through_cycle_rorac_dist"] for s in summaries_6], names_6),
    "n6_win_prob": lambda: win_probability_chart(
        [d["cumulative_profit"] for d in ins_list_6], results_6.config.n_years, names_6),
    "n6_drawdown": lambda: drawdown_chart(
        [d["capital"] for d in ins_list_6], results_6.config.n_years, names_6),
    "n6_buildup": lambda: cumulative_profit_buildup(
        ins_list_6, results_6.config.n_years, names_6),
    "n6_efficiency": lambda: efficiency_frontier(summaries_6, names_6),
    "n6_exec_summary": lambda: executive_summary(
        summaries_6, names_6, 1000, results_6.config.n_years,
        results_6.elapsed_seconds,
        regime_performance_data(results_6.market, ins_list_6, names_6)),
    "n6_ruin": lambda: ruin_over_time_chart(
        [s["ruin_prob_by_year"] for s in summaries_6], results_6.config.n_years, names_6),
    "n6_solvency": lambda: solvency_comparison_chart(
        [d["solvency_ratio"] for d in ins_list_6], results_6.config.n_years, names_6),
}

for name, fn in tests_6.items():
    try:
        fn()
        print(f"  OK: {name}")
    except Exception as e:
        fails.append(name)
        print(f"  FAIL: {name}")
        traceback.print_exc()

# ---- Edge case: n_strategies clamping ----
print(f"\n{'='*50}")
print("EDGE CASE TESTS")
edge_tests = {}

# Test n_strategies out of range gets clamped
from ui.state import build_config
try:
    cfg_clamped = build_config({**params, "n_strategies": 100})
    assert len(cfg_clamped.insurers) == 6, f"Expected 6 (max), got {len(cfg_clamped.insurers)}"
    print("  OK: n_strategies=100 clamped to 6")
except Exception as e:
    fails.append("edge_clamp_high")
    print(f"  FAIL: edge_clamp_high")
    traceback.print_exc()

try:
    cfg_clamped_low = build_config({**params, "n_strategies": 0})
    assert len(cfg_clamped_low.insurers) == 2, f"Expected 2 (min), got {len(cfg_clamped_low.insurers)}"
    print("  OK: n_strategies=0 clamped to 2")
except Exception as e:
    fails.append("edge_clamp_low")
    print(f"  FAIL: edge_clamp_low")
    traceback.print_exc()

# Test Excel export with N>2 strategies
from cyclesim.io import export_results_to_excel
from io import BytesIO
try:
    buf = BytesIO()
    export_results_to_excel(results_6, buf)
    buf.seek(0)
    import openpyxl
    wb = openpyxl.load_workbook(buf, read_only=True)
    for i in range(n6):
        sname = results_6.config.insurers[i].name[:31]
        assert sname in wb.sheetnames, f"Missing sheet '{sname}'"
    wb.close()
    print(f"  OK: Excel export with {n6} strategies ({len(wb.sheetnames)} sheets)")
except Exception as e:
    fails.append("edge_excel_n6")
    print(f"  FAIL: edge_excel_n6")
    traceback.print_exc()

# Test optimizer FullOptResult backward compat with empty gaps
from cyclesim.optimizer import FullOptResult
try:
    fopt_empty = FullOptResult(by_regime={}, unconditional=None, current_gaps=[])
    assert fopt_empty.current_a_gap == {}
    assert fopt_empty.current_b_gap == {}
    fopt_3 = FullOptResult(by_regime={}, unconditional=None,
        current_gaps=[{"x": 1}, {"x": 2}, {"x": 3}])
    assert fopt_3.current_a_gap == {"x": 1}
    assert fopt_3.current_b_gap == {"x": 2}
    print("  OK: FullOptResult backward compat")
except Exception as e:
    fails.append("edge_fullopt_compat")
    print(f"  FAIL: edge_fullopt_compat")
    traceback.print_exc()

print(f"\n{'='*50}")
total = len(tests) + len(tests_3) + len(tests_6) + 4  # +4 edge case tests
if fails:
    print(f"FAILURES: {len(fails)} / {total} -- {', '.join(fails)}")
    sys.exit(1)
else:
    print(f"ALL {total} RENDER TESTS PASSED (N=2, N=3, N=6 + edge cases)")
